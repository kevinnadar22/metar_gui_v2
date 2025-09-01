import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def generate_warning_report(ad_warn_output_path, metar_features_path):
    # Read warnings
    ad_warn_df = pd.read_csv(ad_warn_output_path, dtype={'Issue date/time': str})

    # Read extracted METAR features
    with open(metar_features_path, 'r') as f:
        metar_blocks = f.read().split('\nRow ')[1:]  # Split by each row block
        metar_blocks = ['Row ' + block for block in metar_blocks]

    # Compile regex for TS/TSRA variants
    TSRA_REGEX = re.compile(r'(TSRA|TS|FBL TSRA|MOD TSRA|HVY TSRA|MOD TS|FBL TS|HVY TS)', re.IGNORECASE)

    results = []

    for idx, row in ad_warn_df.iterrows():
        sl_no = idx + 1
        sig_wx = str(row.get('Significant Wx', ''))
        gust_val = str(row.get('Gust', ''))
        wind_dir_fcst = row.get('Wind dir (deg)', None)
        issue_time = str(row.get('Issue date/time', '')).zfill(6)
        station = str(row.get('Station', ''))
        validity_from = str(row.get('Validity from', ''))
        validity_to = str(row.get('Validity To', ''))
        true_false = 0
        remark = ''

        # Check for TS/TSRA and gust in warning
        has_tsra = bool(TSRA_REGEX.search(sig_wx))
        has_gust = bool(re.match(r'\d{2,3}KT', gust_val))

        # Find corresponding METAR block
        metar_block = next((b for b in metar_blocks if f'Row {sl_no}:' in b), None)
        metar_lines = metar_block.split('\n') if metar_block else []

        # If the block says FCST/OBS is OBS, skipping extraction
        if metar_block and 'FCST/OBS is OBS, skipping extraction.' in metar_block:
            true_false = 1
            remark = 'OBS'
            
            # Create separate entries for OBS rows too
            if has_gust and has_tsra:
                # Create TWO separate entries: one for gust, one for thunderstorm
                results.append([sl_no, 'Gust warning', issue_time, true_false, remark, station, validity_from, validity_to])
                results.append([sl_no, 'Thunderstorm warning', issue_time, true_false, remark, station, validity_from, validity_to])
            elif has_gust:
                results.append([sl_no, 'Gust warning', issue_time, true_false, remark, station, validity_from, validity_to])
            elif has_tsra:
                results.append([sl_no, 'Thunderstorm warning', issue_time, true_false, remark, station, validity_from, validity_to])
            continue

        found_gust = False
        found_dir = False
        found_cb = False
        gust_reported = ''
        dir_reported = ''
        cb_reported = ''
        tsra_reported = ''
        cb_cloud_group = ''

        for line in metar_lines:
            # Gust
            gust_match = re.search(r'Gust: (\d+)', line)
            if gust_match:
                metar_gust = int(gust_match.group(1))
                # Wind Dir
                dir_match = re.search(r'Wind Dir: (\d+)', line)
                if dir_match and wind_dir_fcst:
                    metar_dir = int(dir_match.group(1))
                    try:
                        fcst_dir = int(wind_dir_fcst)
                        if abs(metar_dir - fcst_dir) <= 30:
                            found_gust = True
                            found_dir = True
                            gust_reported = f'{metar_gust}KT'
                            dir_reported = f'{metar_dir}'
                    except Exception:
                        pass
            # CB cloud detection from Clouds list
            clouds_match = re.search(r"Clouds: \[(.*?)\]", line)
            if clouds_match:
                clouds_list = [c.strip().strip("'") for c in clouds_match.group(1).split(',')]
                cb_groups = [c for c in clouds_list if 'CB' in c]
                if cb_groups:
                    found_cb = True
                    cb_reported = 'CB'
                    cb_cloud_group = cb_groups[0]  # Take the first CB group found
            # TSRA
            if re.search(r'TSRA', line):
                tsra_reported = 'TSRA'

        # Create separate entries for gust and thunderstorm warnings
        # This prevents double-counting in percentage calculations
        
        if has_gust and has_tsra:
            # Create TWO separate entries: one for gust, one for thunderstorm
            
            # 1. Gust warning entry
            gust_true_false = 0
            gust_remark = ''
            if found_gust and found_dir:
                gust_true_false = 1
                gust_remark = f'Gust {gust_reported} Dir {dir_reported} matched'
                if cb_cloud_group:
                    gust_remark += f' {cb_cloud_group} found'
            else:
                gust_remark = 'No gust/direction mismatch'
            
            results.append([sl_no, 'Gust warning', issue_time, gust_true_false, gust_remark, station, validity_from, validity_to])
            
            # 2. Thunderstorm warning entry
            tsra_true_false = 0
            tsra_remark = ''
            if found_cb:
                tsra_true_false = 1
                tsra_remark = ''
                if cb_cloud_group:
                    tsra_remark += f' {cb_cloud_group} found'
            else:
                tsra_remark = 'Missing CB or direction mismatch'
            
            results.append([sl_no, 'Thunderstorm warning', issue_time, tsra_true_false, tsra_remark, station, validity_from, validity_to])
            
        elif has_gust and not has_tsra:
            # Pure gust warning
            if found_gust and found_dir:
                true_false = 1
                remark = f'Gust {gust_reported} Dir {dir_reported} matched'
                if cb_cloud_group:
                    remark += f' {cb_cloud_group} found'
            else:
                remark = 'No gust/direction mismatch'
            
            results.append([sl_no, 'Gust warning', issue_time, true_false, remark, station, validity_from, validity_to])
            
        elif has_tsra:
            # Pure thunderstorm warning
            if found_cb:
                true_false = 1
                remark = ''
                if cb_cloud_group:
                    remark += f' {cb_cloud_group} found'
            else:
                remark = 'Missing CB or direction mismatch'
            
            results.append([sl_no, 'Thunderstorm warning', issue_time, true_false, remark, station, validity_from, validity_to])

    # Output report
    final_df = pd.DataFrame(results, columns=[
        'Sl. No.',
        'Elements (Thunderstorm/Surface wind & Gust)',
        'Warning issue Time',
        'true-1 / false-0',
        'Remarks',
        'Station',
        'Validity From',
        'Validity To'
    ])
    
    # Add additional columns for better parsing
    final_df['Accuracy_Percentage'] = final_df['true-1 / false-0'].apply(lambda x: f"{x*100}%" if x == 1 else "0%")
    final_df['Warning_Type'] = final_df['Elements (Thunderstorm/Surface wind & Gust)'].apply(
        lambda x: 'Thunderstorm' if 'thunderstorm' in x.lower() else 
                  'Wind' if 'wind' in x.lower() or 'gust' in x.lower() else 'Other'
    )
    
    # Save to a file in the same directory as the input file
    output_path = os.path.join(os.path.dirname(ad_warn_output_path), 'final_warning_report.csv')
    final_df.to_csv(output_path, index=False)
    print('Report saved as final_warning_report.csv')

    # Calculate percentage correct
    total = len(final_df)
    correct = final_df['true-1 / false-0'].sum()
    
    try:
        df_gust = final_df[final_df['Elements (Thunderstorm/Surface wind & Gust)'] == 'Gust warning']
        gust = df_gust['true-1 / false-0'].sum()
        total_gust = len(df_gust)
        df_tsra = final_df[final_df['Elements (Thunderstorm/Surface wind & Gust)'] == 'Thunderstorm warning']
        tsra = df_tsra['true-1 / false-0'].sum()
        total_tsra = len(df_tsra)

        if total > 0:
            percent = (correct / total) * 100
            print(f'Aerodrome Warning : {percent:.0f} % accurate')
        if total_gust > 0:
            percent_gust = (gust / total_gust) * 100
            print(f'Gust warning : {percent_gust:.0f} % accurate')
        if total_tsra > 0:
            percent_tsra = (tsra / total_tsra) * 100
            print(f'Thunderstorm warning : {percent_tsra:.0f} % accurate')
        else:
            print('No warnings to evaluate.')
    except Exception as e:
        print('Could not calculate accuracy:', e)

    # Return the overall accuracy for the API
    accuracy = (correct / total) * 100 if total > 0 else 0
    return final_df, accuracy

def generate_excel_warning_report(ad_warn_output_path, metar_features_path):
    """
    Generate Excel file with the specific format requested:
    1. Tropical cyclone
    2. Thunderstorms
    3. Hail
    4. Snow
    5. Freezing precipitation
    6. Hoar Frost or rime
    7. Dust storm
    8. Sandstorm
    9. Rising sand or dust
    10. Strong surface wind and gusts / Speed
    11. Direction change
    12. Squall
    13. Volcanic ash
    14. Tsunami
    """
    # Read warnings
    ad_warn_df = pd.read_csv(ad_warn_output_path, dtype={'Issue date/time': str})

    # Read extracted METAR features
    with open(metar_features_path, 'r') as f:
        metar_blocks = f.read().split('\nRow ')[1:]  # Split by each row block
        metar_blocks = ['Row ' + block for block in metar_blocks]

    # Compile regex for TS/TSRA variants
    TSRA_REGEX = re.compile(r'(TSRA|TS|FBL TSRA|MOD TSRA|HVY TSRA|MOD TS|FBL TS|HVY TS)', re.IGNORECASE)

    # Predefined warning types
    warning_types = [
        "Tropical cyclone",
        "Thunderstorms", 
        "Hail",
        "Snow",
        "Freezing precipitation",
        "Hoar Frost or rime",
        "Dust storm",
        "Sandstorm",
        "Rising sand or dust",
        "Strong surface wind and gusts / Speed",
        "Direction change",
        "Squall",
        "Volcanic ash",
        "Tsunami"
    ]

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Aerodrome Warning Report"

    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add headers
    headers = ["Sl. No.", "Warning Type", "Issue Time", "Status", "Remarks"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50

    # Analyze the warning data to extract gust and TSRA information
    thunderstorm_data = []
    gust_data = []
    direction_data = []
    
    for idx, row in ad_warn_df.iterrows():
        sl_no = idx + 1
        sig_wx = str(row.get('Significant Wx', ''))
        gust_val = str(row.get('Gust', ''))
        wind_dir_fcst = row.get('Wind dir (deg)', None)
        issue_time = str(row.get('Issue date/time', '')).zfill(6)
        
        # Check for TS/TSRA in warning
        has_tsra = bool(TSRA_REGEX.search(sig_wx))
        has_gust = bool(re.match(r'\d{2,3}KT', gust_val))
        
        if has_tsra:
            thunderstorm_data.append({
                'sl_no': sl_no,
                'issue_time': issue_time,
                'sig_wx': sig_wx
            })
        
        if has_gust:
            gust_data.append({
                'sl_no': sl_no,
                'issue_time': issue_time,
                'gust': gust_val
            })
        
        if wind_dir_fcst is not None:
            direction_data.append({
                'sl_no': sl_no,
                'issue_time': issue_time,
                'direction': wind_dir_fcst
            })

    row_num = 2
    for idx, warning_type in enumerate(warning_types, 1):
        # Initialize default values
        issue_time = ""
        status = "Not Applicable"
        remarks = ""
        
        # Check if this warning type is relevant based on the data
        if warning_type == "Thunderstorms":
            if thunderstorm_data:
                issue_time = thunderstorm_data[0]['issue_time']
                status = "Active"
                remarks = f"Thunderstorm warnings detected: {len(thunderstorm_data)} instances"
                if len(thunderstorm_data) > 1:
                    remarks += f" (First at {thunderstorm_data[0]['issue_time']}, Last at {thunderstorm_data[-1]['issue_time']})"
            else:
                status = "Not Active"
                remarks = "No thunderstorm warnings found in data"
                
        elif warning_type == "Strong surface wind and gusts / Speed":
            if gust_data:
                issue_time = gust_data[0]['issue_time']
                status = "Active"
                remarks = f"Gust warnings detected: {len(gust_data)} instances"
                if len(gust_data) > 1:
                    remarks += f" (First at {gust_data[0]['issue_time']}, Last at {gust_data[-1]['issue_time']})"
                # Add specific gust values
                gust_values = [item['gust'] for item in gust_data]
                remarks += f" - Gust values: {', '.join(gust_values)}"
            else:
                status = "Not Active"
                remarks = "No gust warnings found in data"
                
        elif warning_type == "Direction change":
            if direction_data:
                issue_time = direction_data[0]['issue_time']
                status = "Active"
                remarks = f"Wind direction warnings detected: {len(direction_data)} instances"
                if len(direction_data) > 1:
                    remarks += f" (First at {direction_data[0]['issue_time']}, Last at {direction_data[-1]['issue_time']})"
                # Add specific direction values
                direction_values = [str(item['direction']) for item in direction_data]
                remarks += f" - Directions: {', '.join(direction_values)}"
            else:
                status = "Not Active"
                remarks = "No wind direction warnings found in data"
        
        # Add row to worksheet
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=warning_type).border = border
        ws.cell(row=row_num, column=3, value=issue_time).border = border
        ws.cell(row=row_num, column=4, value=status).border = border
        ws.cell(row=row_num, column=5, value=remarks).border = border
        
        # Apply conditional formatting for status
        status_cell = ws.cell(row=row_num, column=4)
        if status == "Active":
            status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            status_cell.font = Font(color="FFFFFF", bold=True)
        elif status == "Not Active":
            status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            status_cell.font = Font(color="000000", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            status_cell.font = Font(color="000000", bold=True)
        
        row_num += 1

    # Save the Excel file
    output_path = os.path.join(os.path.dirname(ad_warn_output_path), 'aerodrome_warning_report.xlsx')
    wb.save(output_path)
    print(f'Excel report saved as {output_path}')
    
    return output_path 

def generate_aerodrome_warnings_table(ad_warn_output_path, metar_features_path):
    """
    Generate Excel file that matches exactly the frontend table format
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Read the final warning report to get accurate data
    final_report_path = os.path.join(os.path.dirname(ad_warn_output_path), 'final_warning_report.csv')
    
    if not os.path.exists(final_report_path):
        # If final report doesn't exist, generate it first
        generate_warning_report(ad_warn_output_path, metar_features_path)
    
    # Read the final warning report
    with open(final_report_path, 'r', encoding='utf-8') as f:
        heading = f.readline().strip()
    final_df = pd.read_csv(final_report_path, skiprows=1)

    # Create workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Aerodrome Warnings"

    # Add merged, bold heading at the top
    num_columns = 5  # Number of columns in the table
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    heading_cell = ws.cell(row=1, column=1, value=heading)
    heading_cell.font = Font(bold=True, size=14)
    heading_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Define headers exactly as in frontend
    headers = [
        "क्र. सं. / Sr.No.",
        "तत्त्व / Elements", 
        "विमान क्षेत्र चेतावनियों की सं. / Warning no 1 Warning no 2................ Warning no. (Issue date/Issue Time UTC)",
        "रेंज के अंतगर्द आने वाले (पारस) मामलों की प्रतिशतता अथवा सही होने की प्रतिशतता / % of cases within range or occurrence (% correct)",
        "वास्तविक मौसम समयावधि के साथ जिसके लिये चेतावनी जारी नहीं की गयी थी / Actual weather with duration for which no warning was issued"
    ]

    # Add headers to sheet (now row 2)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")

    # Data rows start from row 3
    # Extract thunderstorm and gust data from the final report
    thunderstorm_times = []
    gust_times = []
    
    for _, row in final_df.iterrows():
        elements = str(row.get('Elements (Thunderstorm/Surface wind & Gust)', ''))
        issue_time = str(row.get('Warning issue Time', ''))
        accuracy = row.get('true-1 / false-0', 0)
        
        if 'Thunderstorm' in elements:
            thunderstorm_times.append(issue_time)
        if 'Gust' in elements:
            gust_times.append(issue_time)
    
    # Calculate percentages
    # For thunderstorm: count all entries containing 'Thunderstorm'
    total_thunderstorm = len(final_df[
        final_df['Elements (Thunderstorm/Surface wind & Gust)'].str.contains('Thunderstorm', na=False)
    ])
    
    # For gust: count ONLY pure 'Gust warning' entries (not 'Gust & Thunderstorm warning')
    total_gust = len(final_df[
        final_df['Elements (Thunderstorm/Surface wind & Gust)'] == 'Gust warning'
    ])
    
    # Count accurate predictions (where true-1 / false-0 = 1)
    # For thunderstorm: count all entries containing 'Thunderstorm'
    accurate_thunderstorm = len(final_df[
        (final_df['Elements (Thunderstorm/Surface wind & Gust)'].str.contains('Thunderstorm', na=False)) & 
        (final_df['true-1 / false-0'] == 1)
    ])
    
    # For gust: count ONLY pure 'Gust warning' entries (not 'Gust & Thunderstorm warning')
    accurate_gust = len(final_df[
        (final_df['Elements (Thunderstorm/Surface wind & Gust)'] == 'Gust warning') & 
        (final_df['true-1 / false-0'] == 1)
    ])
    
    thunderstorm_percentage = f"{int((accurate_thunderstorm / total_thunderstorm * 100))}%" if total_thunderstorm > 0 else "0%"
    gust_percentage = f"{int((accurate_gust / total_gust * 100))}%" if total_gust > 0 else "0%"
    
    # Format thunderstorm times - all times separated by commas
    thunderstorm_times_str = ",".join(thunderstorm_times) if thunderstorm_times else "-"
    gust_times_str = ",".join(gust_times) if gust_times else "-"

    # Add data rows exactly as in frontend
    data = [
        ["1.", "उष्णकटिबंधीय चक्रवात / Tropical cyclone", "-", "-", "-"],
        ["2.", "गर्जन सुनामी / Thunderstorms", thunderstorm_times_str, thunderstorm_percentage, "-"],
        ["3.", "ओला / Hail", "-", "-", "-"],
        ["4.", "बर्फ / Snow", "-", "-", "-"],
        ["5.", "हिमवर्षा / Freezing precipitation", "-", "-", "-"],
        ["6.", "पाला या शीत / Hoar Frost or rime", "-", "-", "-"],
        ["7.", "धूल भरी आँधी / Dust storm", "-", "-", "-"],
        ["8.", "रेतीली आँधी / Sandstorm", "-", "-", "-"],
        ["9.", "उठती रेत या धूल / Rising sand or dust", "-", "-", "-"],
        ["10.", "प्रबल सतही पवन तथा झोंके / Strong surface wind and gusts\nगति / Speed", gust_times_str, gust_percentage, "-"],
        ["", "दिशा परिवर्तन / Direction change", "-", "-", "-"],
        ["11.", "बवंडर / Squall\nदिशा / Direction\nगति / Speed", "-", "-", "-"],
        ["12.", "पाला / Frost", "-", "-", "-"],
        ["13.", "ज्वालामुखीय राख / Volcanic ash", "-", "-", "-"],
        ["14.", "सुनामी / Tsunami", "-", "-", "-"],
    ]

    # Add data rows
    for row_idx, row_data in enumerate(data, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40

    # Save Excel file
    output_path = os.path.join(os.path.dirname(ad_warn_output_path), 'Aerodrome_Warnings_Table.xlsx')
    wb.save(output_path)
    print(f'Aerodrome warnings table saved as {output_path}')
    
    return output_path 