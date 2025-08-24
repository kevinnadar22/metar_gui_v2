import requests
from urllib.parse import quote
from app.config import UPPER_AIR_DATA_DIR 
import os
from werkzeug.utils import secure_filename


def fetch_upper_air_data(datetime_str: str, station_id: str, src: str = 'UNKNOWN', data_type: str = 'TEXT:CSV') -> str:
    """
    Fetch upper air sounding data from University of Wyoming's weather site.

    Args:
        datetime_str (str): DateTime in format "YYYY-MM-DD HH:MM:SS"
        station_id (str): 5-digit WMO station ID (e.g. "43003")
        src (str): Source (default is 'UNKNOWN')
        data_type (str): Format type (default is 'TEXT:CSV')

    Returns:
        str: Raw upper air data as plain text

    Raises:
        Exception: If data not available or fetch fails
    """
    base_url = "https://weather.uwyo.edu/wsgi/sounding"
    datetime_encoded = quote(datetime_str)
    full_url = f"{base_url}?datetime={datetime_encoded}&id={station_id}&src={src}&type={data_type}"

    print(f"[DEBUG] Called fetch_upper_air_data with datetime_str={datetime_str}, station_id={station_id}")
    print(f"[DEBUG] Fetching from URL: {full_url}")
    response = requests.get(full_url)

    print(f"[DEBUG] Response status: {response.status_code}")
    print(f"[DEBUG] Response first 100 chars: {response.text[:100]}")

    if response.status_code == 200:
        if '<html>' in response.text.lower():
            raise Exception("HTML page received: likely no data available for this datetime/station.")
        # Save to file
        download_dir = os.path.join(UPPER_AIR_DATA_DIR, 'downloads')
        os.makedirs(download_dir, exist_ok=True)
        dt = datetime_str.replace(":", "").replace("-", "").replace(" ", "_")
        filename = secure_filename(f"upper_air_{station_id}_{dt}.csv")
        file_path = os.path.join(download_dir, filename)
        print(f"[DEBUG] Saving to: {file_path}")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(response.text)
        print(f"[INFO] Data saved to {file_path}")
        return file_path
    else:
        raise Exception(f"Failed to fetch data. HTTP Status Code: {response.status_code}")
    
import pandas as pd
def interpolate_temperature_only(actual_df, forecast_df):
    results = []

    for _, forecast_row in forecast_df.iterrows():
        forecast_alt = forecast_row["Altitude (m)"]

        # Get two actual levels above and below
        below = actual_df[actual_df["geopotential height_m"] <= forecast_alt]
        above = actual_df[actual_df["geopotential height_m"] >= forecast_alt]

        if below.empty or above.empty:
            continue  # Skip if interpolation not possible

        lower = below.iloc[-1]
        upper = above.iloc[0]

        h1, h2 = lower["geopotential height_m"], upper["geopotential height_m"]
        print(f"[DEBUG] Lower level: {h1} m, Upper level: {h2} m for forecast altitude {forecast_alt} m")
        t1, t2 = lower["temperature_C"], upper["temperature_C"]

        # Interpolate temperature
        interp_temp = ((h2 - forecast_alt) * t1 + (forecast_alt - h1) * t2) / (h2 - h1)
        print(f"[DEBUG] Interpolated temperature at {forecast_alt} m: {interp_temp:.2f} C")

        # For other parameters, take the closer one (nearest actual level)
        if abs(h1 - forecast_alt) <= abs(h2 - forecast_alt):
            nearest_row = lower
        else:
            nearest_row = upper

        results.append({
            **forecast_row.to_dict(),
            "interp_temperature_C": interp_temp,
            "actual_wind_speed_m/s": nearest_row["wind speed_m/s"],
            "actual_wind_direction": nearest_row.get("wind direction_degree")
        })

    return pd.DataFrame(results)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# def generate_upper_air_verification_xlsx(data_rows, metadata, file_path,weather_info=None):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Upper Air Verification"

#     thin_border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )

#     def is_merged(ws, row, col):
#         for merged_range in ws.merged_cells.ranges:
#             if (row >= merged_range.min_row and row <= merged_range.max_row and
#                 col >= merged_range.min_col and col <= merged_range.max_col):
#                 return (merged_range.min_row, merged_range.min_col) != (row, col)
#         return False

#     def write_cell(ws, row, col, value, bold=False, center=True):
#         if is_merged(ws, row, col):
#             return  # Skip writing to non-top-left merged cells
#         cell = ws.cell(row=row, column=col)
#         cell.value = value
#         if bold:
#             cell.font = Font(bold=True)
#         if center:
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = thin_border
    

#     # Assign header values BEFORE merging
#     write_cell(ws, 1, 1, "Date", bold=True)
#     write_cell(ws, 1, 2, "Validity (UTC)", bold=True)
#     write_cell(ws, 1, 3, "Forecast & Elements verified", bold=True)
#     write_cell(ws, 1, 7, "Realised & Elements verified", bold=True)
#     write_cell(ws, 1, 10, "Accuracy", bold=True)

#     # Merge header cells
#     ws.merge_cells('A1:A2')
#     ws.merge_cells('B1:B2')
#     ws.merge_cells('C1:F1')
#     ws.merge_cells('G1:I1')
#     ws.merge_cells('J1:L1')

#     # Write bottom row headers (row 2)
#     headers_bottom = [
#         "", "", "FL", "Wind Direction", "Speed", "Temp.",
#         "Wind Direction", "Speed (KT)", "Temp.",
#         "Wind Dir", "Speed", "Temp"
#     ]
#     for col_num, val in enumerate(headers_bottom, start=1):
#         write_cell(ws, 2, col_num, val, bold=True)

#     current_row = 3

#     last_key = None


#     # Write data rows
#     for idx, row in enumerate(data_rows):
#         row_key = f"{row.get('date', '')}_{row.get('validity', '')}"
#         print(f"Checking weather_info for key: {row_key}")
#         print("Available keys in weather_info:", weather_info.keys())


#         values = [
#             row.get("date", ""),
#             row.get("validity", ""),
#             row.get("fl", ""),
#             row.get("forecast_wind_dir", ""),
#             row.get("forecast_speed", ""),
#             row.get("forecast_temp", ""),
#             row.get("actual_wind_dir", ""),
#             row.get("actual_speed", ""),
#             f"{float(row.get('actual_temp', 0)):.2f}",
#             row.get("wind_dir_acc", ""),
#             row.get("speed_acc", ""),
#             row.get("temp_acc", "")
#         ]
#         for col_num, val in enumerate(values, start=1):
#             write_cell(ws, current_row, col_num, val)
#         current_row += 1


#         next_row_key = None
#         if idx + 1 < len(data_rows):
#             next_row = data_rows[idx + 1]
#             next_row_key = f"{next_row.get('date', '')}_{next_row.get('validity', '')}"

#         if row_key != next_row_key:
#             # Add Significant Weather row
#             weather_forecast = weather_info.get(row_key, {}).get("weather_forecast", "") if weather_info else ""
#             weather_realised = " / ".join(weather_info.get(row_key, {}).get("matched", [])) if weather_info else ""
#             weather_accuracy = weather_info.get(row_key, {}).get("accuracy", "") if weather_info else ""
#             print(f"weather_forecast: {weather_forecast}")
#             print(f"weather_realised: {weather_realised}")
#             print(f"weather_accuracy: {weather_accuracy}")

#             weather_row = [
#                 row.get("date", ""),
#                 row.get("validity", ""),
#                 "Significant Weather",
#                 "", "", weather_forecast,
#                 "", "", weather_realised,
#                 "", "", weather_accuracy
#             ]
#             for col_num, val in enumerate(weather_row, start=1):
#                 write_cell(ws, current_row, col_num, val)
#             current_row += 1

#     # Auto adjust column widths
#     for col in ws.columns:
#         max_length = max(len(str(cell.value or "")) for cell in col)
#         ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

#     wb.save(file_path)
#     return file_path

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_upper_air_verification_xlsx(data_rows, metadata, file_path, weather_info=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Upper Air Verification"

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def is_merged(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and
                col >= merged_range.min_col and col <= merged_range.max_col):
                return (merged_range.min_row, merged_range.min_col) != (row, col)
        return False

    def write_cell(ws, row, col, value, bold=False, center=True):
        if is_merged(ws, row, col):
            return
        cell = ws.cell(row=row, column=col)
        cell.value = value
        if bold:
            cell.font = Font(bold=True)
        if center:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Push table down by 2 rows for the heading
    ws.insert_rows(1, amount=2)

    # --- Bilingual Heading ---
    if data_rows:
        first_date = data_rows[0].get("date", "")
        try:
            dt_obj = datetime.strptime(first_date, "%d/%m/%Y")
            month_name = dt_obj.strftime("%B")
            year_val = dt_obj.year
        except:
            month_name = "Unknown"
            year_val = ""

        heading_en = f"Verification of Local / Area Forecasts for the month of {month_name} - {year_val} in r/o AMO Mumbai"
        heading_hi = f"{month_name},{year_val} के लिए लोकल /एरिया पूर्वानुमान का सत्यापन रिपोर्ट"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.cell(row=1, column=1, value=heading_en).font = Font(bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        ws.cell(row=2, column=1, value=heading_hi).font = Font(bold=True)
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # --- Table Header Row (starts from row 3 now) ---
    write_cell(ws, 3, 1, "Date", bold=True)
    write_cell(ws, 3, 2, "Validity (UTC)", bold=True)
    write_cell(ws, 3, 3, "Forecast & Elements verified", bold=True)
    write_cell(ws, 3, 7, "Realised & Elements verified", bold=True)
    write_cell(ws, 3, 10, "Accuracy", bold=True)

    # Merge header cells
    ws.merge_cells('A3:A4')
    ws.merge_cells('B3:B4')
    ws.merge_cells('C3:F3')
    ws.merge_cells('G3:I3')
    ws.merge_cells('J3:L3')

    # Row 4 headers
    headers_bottom = [
        "", "", "FL", "Wind Direction", "Speed", "Temp.",
        "Wind Direction", "Speed (KT)", "Temp.",
        "Wind Dir", "Speed", "Temp"
    ]
    for col_num, val in enumerate(headers_bottom, start=1):
        write_cell(ws, 4, col_num, val, bold=True)

    # --- Data Rows start at row 5 ---
    current_row = 5

    altitude_to_fl = {
    3000: "FL 100 (3000 M)",
    2100: "FL 070 (2100 M)",
    1500: "FL 050 (1500 M)",
    900:  "FL 030 (900 M)",
    600:  "FL 020 (600 M)",
    300:  "FL 010 (300 M)"
    }

    fl_labels = list(altitude_to_fl.values())
    filtered_rows = [row for row in data_rows if row.get("fl") in fl_labels]

    overall_temp_acc = round(sum(1 for r in data_rows if r.get("temp_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0
    overall_wind_acc = round(sum(1 for r in data_rows if r.get("speed_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0
    overall_wind_dir_acc = round(sum(1 for r in data_rows if r.get("wind_dir_acc") == "CORRECT") / len(data_rows) * 100, 2) if data_rows else 0

# For FL accuracy (only FL levels)
    fl_temp_acc = round(sum(1 for r in filtered_rows if r.get("temp_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None
    fl_wind_acc = round(sum(1 for r in filtered_rows if r.get("speed_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None
    fl_wind_dir_acc = round(sum(1 for r in filtered_rows if r.get("wind_dir_acc") == "CORRECT") / len(filtered_rows) * 100, 2) if filtered_rows else None


    for idx, row in enumerate(filtered_rows):
        row_key = f"{row.get('date', '')}_{row.get('validity', '')}"
        print(f"Checking weather_info for key: {row_key}")
        if weather_info:
            print("Available keys in weather_info:", weather_info.keys())

        values = [
            row.get("date", ""),
            row.get("validity", ""),
            row.get("fl", ""),
            row.get("forecast_wind_dir", ""),
            row.get("forecast_speed", ""),
            row.get("forecast_temp", ""),
            row.get("actual_wind_dir", ""),
            row.get("actual_speed", ""),
            f"{float(row.get('actual_temp', 0)):.2f}",
            row.get("wind_dir_acc", ""),
            row.get("speed_acc", ""),
            row.get("temp_acc", ""),
            # row.get("fl_accuracy_summary", {}),
        ]
        for col_num, val in enumerate(values, start=1):
            write_cell(ws, current_row, col_num, val)
        current_row += 1
        print(values)

        # Check next row key
        next_row_key = None
        if idx + 1 < len(data_rows):
            next_row = data_rows[idx + 1]
            next_row_key = f"{next_row.get('date', '')}_{next_row.get('validity', '')}"

        if row_key != next_row_key or idx == len(filtered_rows) - 1:
            weather_forecast = weather_info.get(row_key, {}).get("weather_forecast", "") if weather_info else ""
            weather_realised = " / ".join(weather_info.get(row_key, {}).get("matched", [])) if weather_info else ""
            weather_accuracy = weather_info.get(row_key, {}).get("accuracy", "") if weather_info else ""

            print(f"Writing weather row for {row_key}")  # DEBUG
            weather_row = [
                row.get("date", ""),
                row.get("validity", ""),
                "Significant Weather",
                "", "", weather_forecast,
                "", "", weather_realised,
                "", "", weather_accuracy
            ]
            for col_num, val in enumerate(weather_row, start=1):
                write_cell(ws, current_row, col_num, val)
            current_row += 1

    
    current_row += 1
  # or filter differently if needed

# Calculate overall accuracies
    ws.cell(row=current_row, column=1, value=f"Overall Temperature Accuracy: {overall_temp_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value=f"Overall Wind Speed Accuracy: {overall_wind_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

    ws.cell(row=current_row, column=1, value=f"Overall Wind Direction Accuracy: {overall_wind_dir_acc}%")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="Accuracy Summary for FL 010 (300 M) to FL 100 (3000 M)")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1

# Table header
    headers = ["FL Range", "Temperature Accuracy (%)", "Wind Speed Accuracy (%)", "Wind Direction Accuracy (%)"]
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=current_row, column=col_num, value=header)
        ws.cell(row=current_row, column=col_num).font = Font(bold=True)
    current_row += 1

# Single summary row for all FLs
    ws.cell(row=current_row, column=1, value="FL 010 (300 M) to FL 100 (3000 M)")
    ws.cell(row=current_row, column=2, value=fl_temp_acc)
    ws.cell(row=current_row, column=3, value=fl_wind_acc)
    ws.cell(row=current_row, column=4, value=fl_wind_dir_acc)
    # ws.cell(row=current_row, column=5, value=fl_accuracy_summary["Total Levels"])
    current_row += 1

    # Leave a blank row after FL summary
    

    # Auto adjust widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    wb.save(file_path)
    return file_path



# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# import os

# def generate_upper_air_verification_xlsx(data_rows, metadata, file_path):
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Upper Air Verification"

#     thin_border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )

#     def write_cell(row, col, value, bold=False, center=True):
#         cell = ws.cell(row=row, column=col)
#         for merged_range in ws.merged_cells.ranges:
#             if (row, col) in merged_range and (row, col) != (merged_range.min_row, merged_range.min_col):
#                 return
#         cell.value = value
#         if bold:
#             cell.font = Font(bold=True)
#         if center:
#             cell.alignment = Alignment(horizontal="center", vertical="center")
#         cell.border = thin_border

#     # Assign header values BEFORE merging
#     write_cell(1, 1, "Date", bold=True)
#     write_cell(1, 2, "Validity (UTC)", bold=True)
#     write_cell(1, 3, "Forecast & Elements verified", bold=True)
#     write_cell(1, 7, "Realised & Elements verified", bold=True)
#     write_cell(1, 10, "Accuracy", bold=True)

#     # Merge header cells
#     ws.merge_cells('A1:A2')
#     ws.merge_cells('B1:B2')
#     ws.merge_cells('C1:F1')
#     ws.merge_cells('G1:I1')
#     ws.merge_cells('J1:L1')

#     headers_bottom = [
#         "", "", "FL", "Wind Direction", "Speed", "Temp.",
#         "Wind Direction", "Speed (KT)", "Temp.",
#         "Wind Dir", "Speed", "Temp"
#     ]
#     for col_num, val in enumerate(headers_bottom, start=1):
#         write_cell(2, col_num, val, bold=True)

    

#     for idx, row in enumerate(data_rows, start=3):
        
#         values = [
#             row.get("date", ""),
#             row.get("validity", ""),
            
#             row.get("forecast_wind_dir", ""),
#             row.get("forecast_speed", ""),
#             f"{float(row.get('forecast_temp', 0)):.2f}",
#             row.get("actual_wind_dir", ""),
#             round(float(row.get("actual_speed", 0)), 2),
#             f"{float(row.get('actual_temp', 0)):.2f}",
#             row.get("wind_dir_acc", ""),
#             row.get("speed_acc", ""),
#             row.get("temp_acc", "")
#         ]
#         for col_num, val in enumerate(values, start=1):
#             write_cell(idx, col_num, val)

#     for col in ws.columns:
#         max_length = max(len(str(cell.value or "")) for cell in col)
#         ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

#     wb.save(file_path)
#     return file_path

